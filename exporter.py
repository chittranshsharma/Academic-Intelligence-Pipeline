import json
import logging
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column display widths (in Excel units). Adjust as needed.
COLUMN_WIDTHS = {
    "S No":           6,
    "Region":         10,
    "University Name": 30,
    "Department":     28,
    "Faculty Name":   22,
    "Origin":         14,
    "Position":       36,
    "Email":          30,
    "Phone":          18,
    "Profile link":   45,
    "Research":       50,
    "Notes":          60,
}

HEADER_COLOR  = "1F3864"   # Dark navy
HEADER_FONT   = "FFFFFF"   # White text
ALT_ROW_COLOR = "DCE6F1"   # Light blue alternating rows


class FacultyExporter:
    def __init__(self, input_json="cleaned_data.json", output_dir="output"):
        self.input_json = input_json
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.csv_path  = os.path.join(self.output_dir, "faculty_data.csv")
        self.xlsx_path = os.path.join(self.output_dir, "faculty_data.xlsx")

    def _build_dataframe(self, data):
        df = pd.DataFrame(data)

        # ── Deduplicate ──
        if "profile_link" in df.columns:
            before = len(df)
            df = df.drop_duplicates(subset=["profile_link"], keep="first")
            removed = before - len(df)
            if removed:
                logger.info(f"Removed {removed} duplicate entries.")

        # ── Clean phone / profile_link ──
        if "phone" in df.columns:
            df["phone"] = df["phone"].fillna("NA").replace("", "NA")
        if "profile_link" in df.columns:
            df["profile_link"] = df["profile_link"].fillna("")

        # ── Rename columns ──
        rename_mapping = {
            "country":          "Region",
            "university":       "University Name",
            "department":       "Department",
            "name":             "Faculty Name",
            "origin":           "Origin",
            "role":             "Position",
            "email":            "Email",
            "phone":            "Phone",
            "profile_link":     "Profile link",
            "research_interests": "Research",
            "summary":          "Notes",
        }
        df = df.rename(columns=rename_mapping)

        # ── Column order ──
        ordered_cols = [
            "Region", "University Name", "Department", "Faculty Name",
            "Origin", "Position", "Email", "Phone", "Profile link",
            "Research", "Notes",
        ]
        existing = [c for c in ordered_cols if c in df.columns]
        df = df[existing]

        # ── Fill any remaining NaN in text cols ──
        df = df.fillna("")

        # ── S No ──
        df.insert(0, "S No", range(1, len(df) + 1))
        return df

    def _style_xlsx(self, path):
        """Apply professional styling: bold header, column widths, alternating rows, wrapping."""
        wb = load_workbook(path)
        ws = wb.active

        header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
        alt_fill    = PatternFill("solid", fgColor=ALT_ROW_COLOR)
        thin_border = Border(
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
            top=Side(style="thin", color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"),
        )

        # Style header row
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font      = Font(bold=True, color=HEADER_FONT, name="Calibri", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border    = thin_border

            # Set column width from our map, fall back to auto-detect
            col_name = cell.value or ""
            width = COLUMN_WIDTHS.get(col_name, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

